# -*- coding: UTF-8 -*-

import os
import shutil
from optparse import make_option

from django.conf import settings
from django.utils.encoding import smart_str
from django.core.management.base import NoArgsCommand
from django.db import models

SILENT, NORMAL, VERBOSE, VERY_VERBOSE = 0, 1, 2, 3


class Command(NoArgsCommand):
    option_list = NoArgsCommand.option_list
    help = "Exports the database to XLSX Workbook"

    def handle_noargs(self, *args, **options):
        import os
        import openpyxl
        from openpyxl.cell import get_column_letter
        self.verbosity = int(options.get("verbosity", NORMAL))

        Location = models.get_model("locations", "Location")
        Stage = models.get_model("locations", "Stage")
        LocationImage = models.get_model("locations", "Image")
        FileDescription = models.get_model("filebrowser", "FileDescription")

        ProductionCategory = models.get_model("productions", "ProductionCategory")
        Production = models.get_model("productions", "Production")
        Event = models.get_model("productions", "Event")
        ProductionImage = models.get_model("productions", "ProductionImage")
        EventImage = models.get_model("productions", "EventImage")

        wb = openpyxl.Workbook()  # encoding='utf-8'

        if self.verbosity >= NORMAL:
            print u"=== Exporting Locations ==="
        ws = wb.active
        ws.title = "Locations"
        ws.append(['id', 'title_de', 'subtitle_de', 'street_address', 'street_address2', 'postal_code', 'city', 'stages', 'images'])
        loc_count = Location.objects.filter(status="published").count()
        for loc_index, loc in enumerate(Location.objects.filter(status="published"), 1):
            print "%d/%d %s" % (loc_index, loc_count, smart_str(loc.title_de))
            ws.append([loc.id, loc.title_de, loc.subtitle_de, loc.street_address, loc.street_address2, loc.postal_code, loc.city,
                u','.join([str(st.pk) for st in loc.stage_set.all()]),
                u','.join([str(im.pk) for im in loc.image_set.all()]),
            ])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Stages ==="
        ws = wb.create_sheet(1, "Stages")
        ws.append(['id', 'location_id', 'title_de', 'street_address', 'street_address2', 'postal_code', 'city'])
        st_count = Stage.objects.filter(location__status="published").distinct().count()
        for st_index, st in enumerate(Stage.objects.filter(location__status="published").distinct(), 1):
            print "%d/%d %s" % (st_index, st_count, smart_str(st.title_de))
            ws.append([st.id, st.location_id, st.title_de, st.street_address, st.street_address2, st.postal_code, st.city])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Location Images ==="
        ws = wb.create_sheet(2, "Location Images")
        ws.append(['id', 'location_id', 'url', 'title_de', 'description_de', 'author', 'copyright_restrictions'])
        im_count = LocationImage.objects.filter(location__status="published").distinct().count()
        for im_index, im in enumerate(LocationImage.objects.filter(location__status="published").distinct(), 1):
            print "%d/%d %s" % (im_index, im_count, smart_str(im.path.path))
            file_desc = FileDescription()
            file_descs = FileDescription.objects.filter(file_path=im.path)
            if file_descs:
                file_desc = file_descs[0]
            ws.append([im.id, im.location_id, 'http://www.berlin-buehnen.de' + settings.MEDIA_URL + im.path.path, file_desc.title_de, file_desc.description_de, file_desc.author, im.copyright_restrictions])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Production Categories ==="
        ws = wb.create_sheet(3, "Production Categories")
        ws.append(['id', 'parent_id', 'title_de'])
        cat_count = ProductionCategory.objects.count()
        for cat_index, cat in enumerate(ProductionCategory.objects.all(), 1):
            print "%d/%d %s" % (cat_index, cat_count, smart_str(cat.title_de))
            ws.append([cat.id, cat.parent_id, cat.title_de])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Productions ==="
        ws = wb.create_sheet(4, "Productions")
        ws.append(['id', 'creation_date', 'modified_date', 'title_de', 'subtitle_de', 'link_de', 'contents_de', 'credits_de', 'concert_program_de', 'supporting_program_de', 'remarks_de', 'subtitles_text_de', 'age_text_de', 'free_entrance', 'in_program_of', 'play_locations', 'play_stages', 'location_title', 'street_address', 'street_address2', 'postal_code', 'city', 'categories', 'leaders', 'authors', 'participants', 'images', 'events',])
        prod_count = Production.objects.filter(status="published").count()
        for prod_index, prod in enumerate(Production.objects.filter(status="published"), 1):
            print "%d/%d %s" % (prod_index, prod_count, smart_str(prod.title_de))
            ws.append([prod.id, prod.creation_date, prod.modified_date, prod.title_de, prod.subtitle_de, prod.get_url(), prod.get_rendered_contents_de(), prod.get_rendered_credits_de(), prod.get_rendered_concert_program_de(), prod.get_rendered_supporting_program_de(), prod.get_rendered_remarks_de(), prod.subtitles_text_de, prod.age_text_de, prod.free_entrance,
                u','.join(map(str, prod.in_program_of.values_list('id', flat=True))),
                u','.join(map(str, prod.play_locations.values_list('id', flat=True))),
                u','.join(map(str, prod.play_stages.values_list('id', flat=True))),
                prod.location_title, prod.street_address, prod.street_address2, prod.postal_code, prod.city,
                u','.join(map(str, prod.categories.values_list('id', flat=True))),
                u','.join([u'%s-%s' % (leader.person, leader.get_function()) for leader in prod.productionleadership_set.all()]),
                u','.join([u'%s-%s' % (author.person, author.get_function()) for author in prod.productionauthorship_set.all()]),
                u','.join([u'%s-%s' % (participant.person, participant.get_function()) for participant in prod.productioninvolvement_set.all()]),
                u','.join([str(im.pk) for im in prod.productionimage_set.all()]),
                u','.join([str(ev.pk) for ev in prod.event_set.all()]),
            ])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Events ==="
        ws = wb.create_sheet(5, "Events")
        ws.append(['id', 'production_id', 'creation_date', 'modified_date', 'link_de', 'start_date', 'end_date', 'start_time', 'play_locations', 'play_stages', 'location_title', 'street_address', 'street_address2', 'postal_code', 'city', 'free_entrance', 'images'])
        ev_count = Event.objects.filter(production__status="published").distinct().count()
        for ev_index, ev in enumerate(Event.objects.filter(production__status="published").distinct(), 1):
            print "%d/%d %s %s" % (ev_index, ev_count, smart_str(ev.start_date), smart_str(ev.start_time))
            ws.append([ev.id, ev.production_id, ev.creation_date, ev.modified_date, ev.get_url(), ev.start_date, ev.end_date, ev.start_time,
                u','.join(map(str, ev.play_locations.values_list('id', flat=True))),
                u','.join(map(str, ev.play_stages.values_list('id', flat=True))),
                ev.location_title, ev.street_address, ev.street_address2, ev.postal_code, ev.city, ev.free_entrance,
                u','.join([str(im.pk) for im in ev.eventimage_set.all()]),
            ])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Production Images ==="
        ws = wb.create_sheet(6, "Production Images")
        ws.append(['id', 'production_id', 'url', 'title_de', 'description_de', 'author', 'copyright_restrictions'])
        im_count = ProductionImage.objects.filter(production__status="published").distinct().count()
        for im_index, im in enumerate(ProductionImage.objects.filter(production__status="published").distinct(), 1):
            print "%d/%d %s" % (im_index, im_count, smart_str(im.path.path))
            file_desc = FileDescription()
            file_descs = FileDescription.objects.filter(file_path=im.path)
            if file_descs:
                file_desc = file_descs[0]
            ws.append([im.id, im.production_id, 'http://www.berlin-buehnen.de' + settings.MEDIA_URL + im.path.path, file_desc.title_de, file_desc.description_de, file_desc.author, im.copyright_restrictions])

        if self.verbosity >= NORMAL:
            print u"=== Exporting Event Images ==="
        ws = wb.create_sheet(7, "Event Images")
        ws.append(['id', 'event_id', 'url', 'title_de', 'description_de', 'author', 'copyright_restrictions'])
        im_count = EventImage.objects.filter(event__production__status="published").distinct().count()
        for im_index, im in enumerate(EventImage.objects.filter(event__production__status="published").distinct(), 1):
            print "%d/%d %s" % (im_index, im_count, smart_str(im.path.path))
            file_desc = FileDescription()
            file_descs = FileDescription.objects.filter(file_path=im.path)
            if file_descs:
                file_desc = file_descs[0]
            ws.append([im.id, im.event_id, 'http://www.berlin-buehnen.de' + settings.MEDIA_URL + im.path.path, file_desc.title_de, file_desc.description_de, file_desc.author, im.copyright_restrictions])

        wb.save(os.path.join(settings.MEDIA_ROOT, 'data-export.xlsx'))
